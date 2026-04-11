import asyncio
import hashlib
import json
import os
import time
import uuid

from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func as sa_func
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from app.api.deps import get_current_user, get_db, require_permission
from app.models.dataset import Dataset, DatasetVersion, SourceType, SyncLog
from app.models.eval_result import EvalResult
from app.models.user import User
from app.schemas.dataset import (
    DatasetImportRequest,
    DatasetMountRequest,
    DatasetResponse,
    DatasetStatsResponse,
    DatasetSubscribeRequest,
    DatasetVersionResponse,
    PaginatedResponse,
    PreflightConfirmRequest,
    PreflightResponse,
    SyncLogResponse,
)
from app.services.dataset_deletion import cleanup_uploaded_file, delete_dataset_versions
from app.services.storage import StorageBackend, get_storage
from app.services.storage.file_io import read_bytes as _read_bytes_raw
from app.services.storage.file_io import read_text as _read_text_raw
from app.services.storage.utils import uri_to_key

# In-memory preflight cache: token -> {source_uri, format, row_count, ...}
_preflight_cache: dict[str, dict] = {}
_PREFLIGHT_TTL = 1800  # 30 minutes
_PREFLIGHT_MAX_SIZE = 200


def _cleanup_preflight_cache() -> None:
    """Remove expired preflight entries."""
    now = time.time()
    expired = [
        k for k, v in _preflight_cache.items() if now - v.get("created_at", 0) > _PREFLIGHT_TTL
    ]
    for k in expired:
        _preflight_cache.pop(k, None)


router = APIRouter()


def _get_storage() -> StorageBackend:
    return get_storage()


def _resolve_path(uri: str) -> str:
    """Get the effective file extension from a URI or key."""
    key = uri_to_key(uri)
    return key if key is not None else uri


async def _count_rows(storage: StorageBackend, source_uri: str) -> int:
    """Count rows in a data file via the storage backend."""
    path = _resolve_path(source_uri)
    lower = path.lower()

    # Parquet — read row count from metadata (no data parsing)
    if lower.endswith(".parquet"):
        content = await _read_bytes(storage, source_uri)
        if content is None:
            return 0
        import io

        import pyarrow.parquet as pq

        return pq.ParquetFile(io.BytesIO(content)).metadata.num_rows

    # Excel — use openpyxl read-only mode for row count
    if lower.endswith((".xlsx", ".xls")):
        content = await _read_bytes(storage, source_uri)
        if content is None:
            return 0
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True)
        count = wb.active.max_row - 1 if wb.active else 0  # minus header
        wb.close()
        return max(0, count)

    # CSV
    if lower.endswith(".csv"):
        text = await _read_text(storage, source_uri)
        if text is None:
            return 0
        return max(0, text.count("\n") - 1)  # minus header

    # JSON
    if lower.endswith(".json"):
        text = await _read_text(storage, source_uri)
        if text is None:
            return 0
        data = json.loads(text)
        return len(data) if isinstance(data, list) else 1

    # JSONL (default)
    text = await _read_text(storage, source_uri)
    if text is None:
        return 0
    return sum(1 for line in text.splitlines() if line.strip())


async def _read_text(storage: StorageBackend, source_uri: str) -> str | None:
    """Read file content as text, from storage or local filesystem."""
    try:
        return await _read_text_raw(storage, source_uri)
    except FileNotFoundError:
        return None


async def _read_bytes(storage: StorageBackend, source_uri: str) -> bytes | None:
    """Read file content as bytes, from storage or local filesystem."""
    try:
        return await _read_bytes_raw(storage, source_uri)
    except FileNotFoundError:
        return None


@router.post("/upload", response_model=DatasetResponse, status_code=201)
async def upload_dataset(
    file: UploadFile,
    name: str = Form(""),
    description: str = Form(""),
    tags: str = Form(""),
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
    storage: StorageBackend = Depends(_get_storage),
):
    if not name:
        name = file.filename or "untitled"

    ext = os.path.splitext(file.filename or "")[1].lower()
    fmt = ext.lstrip(".") or "jsonl"

    content = await file.read()
    file_id = uuid.uuid4()
    key = f"uploads/{file_id}{ext}"
    uri = await storage.write_file(key, content)

    row_count = await _count_rows(storage, uri)

    # Check if dataset with same name exists (auto-version)
    stmt = select(Dataset).where(Dataset.name == name).order_by(Dataset.version.desc())
    existing = (await session.exec(stmt)).first()
    version = (existing.version + 1) if existing else 1

    ds = Dataset(
        name=name,
        description=description,
        source_type=SourceType.upload,
        source_uri=uri,
        format=fmt,
        tags=tags,
        version=version,
        size_bytes=len(content),
        row_count=row_count,
        created_by=current_user.id,
    )
    session.add(ds)
    await session.commit()
    await session.refresh(ds)

    # Create version record
    dv = DatasetVersion(
        dataset_id=ds.id,
        version=version,
        file_path=uri,
        changelog="Initial upload" if version == 1 else f"Version {version}",
        row_count=row_count,
        size_bytes=len(content),
        format=fmt,
    )
    session.add(dv)
    await session.commit()
    await session.refresh(ds)

    return ds


@router.post("/import", response_model=DatasetResponse, status_code=201)
async def import_dataset(
    body: DatasetImportRequest,
    job_id: str = Query("", description="Optional job ID for progress tracking"),
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
    storage: StorageBackend = Depends(_get_storage),
):
    """Import dataset from HuggingFace or ModelScope."""
    from app.services.dataset_import import import_huggingface, import_modelscope
    from app.services.import_progress import create_job, update_job

    if body.source not in ("huggingface", "modelscope"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "source 必须为 huggingface 或 modelscope")

    source_type = SourceType.huggingface if body.source == "huggingface" else SourceType.modelscope
    display_name = body.name or body.dataset_id.split("/")[-1]

    # Create progress tracking job if ID provided
    tracking_id = job_id or str(uuid.uuid4())
    create_job(tracking_id, display_name)
    update_job(tracking_id, status="downloading", phase="开始导入", progress=0.01)

    try:
        if body.source == "huggingface":
            source_uri, row_count, size_bytes = await import_huggingface(
                body.dataset_id,
                body.subset,
                body.split,
                storage,
                job_id=tracking_id,
                hf_token=current_user.hf_token or None,
            )
        else:
            source_uri, row_count, size_bytes = await import_modelscope(
                body.dataset_id,
                body.subset,
                body.split,
                storage,
                ms_token=current_user.ms_token or None,
            )
        update_job(tracking_id, status="done", progress=1.0, phase="完成")
        from app.metrics import dataset_imports_total

        dataset_imports_total.labels(source=body.source, status="success").inc()
    except ValueError as e:
        from app.metrics import dataset_imports_total

        dataset_imports_total.labels(source=body.source, status="error").inc()
        error_msg = str(e)
        lower_msg = error_msg.lower()
        if "401" in error_msg or "unauthorized" in lower_msg or "gated" in lower_msg:
            if not current_user.hf_token and body.source == "huggingface":
                error_msg += "\n提示：请在账号设置中配置 HuggingFace Token"
            elif not current_user.ms_token and body.source == "modelscope":
                error_msg += "\n提示：请在账号设置中配置 ModelScope Token"
        update_job(tracking_id, status="failed", error=error_msg)
        raise HTTPException(status.HTTP_400_BAD_REQUEST, error_msg) from e

    # Auto-version if same name exists
    stmt = select(Dataset).where(Dataset.name == display_name).order_by(Dataset.version.desc())
    existing = (await session.exec(stmt)).first()
    version = (existing.version + 1) if existing else 1

    ext = os.path.splitext(source_uri)[1].lstrip(".")
    ds = Dataset(
        name=display_name,
        description=body.description,
        source_type=source_type,
        source_uri=source_uri,
        format=ext or "jsonl",
        tags=body.tags,
        version=version,
        size_bytes=size_bytes,
        row_count=row_count,
        created_by=current_user.id,
        hf_dataset_id=body.dataset_id,
        hf_subset=body.subset,
        hf_split=body.split,
    )
    session.add(ds)
    await session.commit()
    await session.refresh(ds)

    dv = DatasetVersion(
        dataset_id=ds.id,
        version=version,
        file_path=source_uri,
        changelog=f"Imported from {body.source}: {body.dataset_id}",
        row_count=row_count,
        size_bytes=size_bytes,
        format=ext or "jsonl",
    )
    session.add(dv)
    await session.commit()
    await session.refresh(ds)
    return ds


@router.post("/mount", response_model=DatasetResponse, status_code=201)
async def mount_dataset(
    body: DatasetMountRequest,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
    storage: StorageBackend = Depends(_get_storage),
):
    # Mount always operates on local filesystem paths
    if not os.path.exists(body.server_path):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "服务器路径不存在")

    row_count = await _count_rows(storage, body.server_path)
    size = os.path.getsize(body.server_path)

    ds = Dataset(
        name=body.name,
        description=body.description,
        source_type=SourceType.server_path,
        source_uri=body.server_path,
        format=body.format,
        tags=body.tags,
        version=1,
        size_bytes=size,
        row_count=row_count,
        created_by=current_user.id,
    )
    session.add(ds)
    await session.commit()
    await session.refresh(ds)
    return ds


@router.get("/presets")
async def list_preset_datasets(
    current_user: User = require_permission("datasets.read"),
):
    """Return the catalog of available preset datasets (not stored in DB)."""
    from app.database import PRESET_DATASETS

    return PRESET_DATASETS


@router.get("/import-progress/{job_id}")
async def stream_import_progress(
    job_id: str,
    current_user: User = Depends(get_current_user),
):
    """SSE endpoint streaming import progress for a given job."""
    from app.services.import_progress import get_event, get_job

    async def event_stream():
        while True:
            job = get_job(job_id)
            if not job:
                payload = json.dumps(dict(status="not_found"))
                yield "data: " + payload + "\n\n"
                return
            payload = json.dumps(
                dict(
                    status=job.status,
                    phase=job.phase,
                    progress=job.progress,
                    error=job.error,
                )
            )
            yield "data: " + payload + "\n\n"
            if job.status in ("done", "failed"):
                return
            evt = get_event(job_id)
            if evt:
                try:
                    await asyncio.wait_for(evt.wait(), timeout=2.0)
                except asyncio.TimeoutError:
                    pass

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("", response_model=PaginatedResponse)
async def list_datasets(
    page: int = 1,
    page_size: int = 20,
    tag: str | None = None,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
):
    base = select(Dataset)
    if tag:
        base = base.where(Dataset.tags.contains(tag))

    count_stmt = select(sa_func.count()).select_from(base.subquery())
    total = (await session.exec(count_stmt)).one()

    offset = (page - 1) * page_size
    items_stmt = base.order_by(Dataset.created_at.desc()).offset(offset).limit(page_size)
    result = await session.exec(items_stmt)
    return PaginatedResponse(items=result.all(), total=total, page=page, page_size=page_size)


@router.post("/{dataset_id}/download", response_model=DatasetResponse)
async def download_preset_content(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.download"),
    storage: StorageBackend = Depends(_get_storage),
):
    """Download content for a preset dataset from HuggingFace."""
    from app.database import PRESET_DATASETS
    from app.services.dataset_import import import_huggingface

    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")

    if ds.row_count > 0 and ds.size_bytes > 0:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "数据集已有内容")

    # Resolve HF ID and subset — for presets, look up in PRESET_DATASETS
    hf_id = ""
    subset = ""
    split = "test"
    if ds.source_type == SourceType.preset:
        for preset in PRESET_DATASETS:
            if preset["name"] == ds.name:
                hf_id = preset.get("source_id", preset.get("hf_id", ""))
                subset = preset.get("subset", "")
                split = preset.get("split", "test")
                break
        if not hf_id:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"预设数据集 '{ds.name}' 未在目录中找到",
            )
    elif ds.source_type == SourceType.huggingface:
        # Use stored HF dataset ID, not source_uri (which is local path after import)
        hf_id = ds.hf_dataset_id or ds.source_uri
        subset = ds.hf_subset
        split = ds.hf_split or "test"
    else:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "仅支持预设和 HuggingFace 数据集的下载",
        )

    try:
        source_uri, row_count, size_bytes = await import_huggingface(
            hf_id,
            subset,
            split,
            storage,
            hf_token=current_user.hf_token or None,
        )
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"下载失败: {e}") from e

    ds.source_uri = source_uri
    # After downloading, mark as huggingface (no longer a placeholder)
    if ds.source_type == SourceType.preset:
        ds.source_type = SourceType.huggingface
        if not ds.hf_dataset_id:
            ds.hf_dataset_id = hf_id
    ds.row_count = row_count
    ds.size_bytes = size_bytes
    ext = os.path.splitext(source_uri)[1].lstrip(".")
    if ext:
        ds.format = ext
    session.add(ds)

    # Also create a version record
    dv = DatasetVersion(
        dataset_id=ds.id,
        version=ds.version,
        file_path=source_uri,
        changelog=f"Downloaded from HuggingFace: {hf_id}",
        row_count=row_count,
    )
    session.add(dv)
    await session.commit()
    await session.refresh(ds)
    return ds


@router.get("/{dataset_id}", response_model=DatasetResponse)
async def get_dataset(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
):
    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")
    return ds


@router.get("/{dataset_id}/preview")
async def preview_dataset(
    dataset_id: uuid.UUID,
    limit: int = 50,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
    storage: StorageBackend = Depends(_get_storage),
):
    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")

    import pandas as pd

    path = _resolve_path(ds.source_uri)
    lower = path.lower()

    rows: list[dict] = []

    # Binary formats — Parquet, Excel (read only needed rows)
    if lower.endswith((".parquet", ".xlsx", ".xls")):
        content = await _read_bytes(storage, ds.source_uri)
        if content is None:
            return {"rows": [], "total": 0}
        import io

        if lower.endswith(".parquet"):
            import pyarrow.parquet as pq

            tbl = pq.read_table(io.BytesIO(content)).slice(0, limit)
            df = tbl.to_pandas()
        else:
            df = pd.read_excel(
                io.BytesIO(content),
                nrows=limit,
            )
        rows = df.fillna("").to_dict(orient="records")
        return {"rows": rows, "total": ds.row_count}

    # Text formats — JSON, JSONL, CSV
    text = await _read_text(storage, ds.source_uri)
    if text is None:
        return {"rows": [], "total": 0}

    if lower.endswith(".csv"):
        import io

        df = pd.read_csv(io.StringIO(text), nrows=limit)
        rows = df.fillna("").to_dict(orient="records")
    elif lower.endswith(".json"):
        data = json.loads(text)
        items = data if isinstance(data, list) else [data]
        rows = items[:limit]
    else:
        # JSONL
        for line in text.splitlines():
            if len(rows) >= limit:
                break
            line = line.strip()
            if line:
                rows.append(json.loads(line))

    return {"rows": rows, "total": ds.row_count}


@router.post("/{dataset_id}/subscribe", response_model=DatasetResponse)
async def subscribe_dataset(
    dataset_id: uuid.UUID,
    body: DatasetSubscribeRequest,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
):
    """Enable auto-update subscription for a dataset."""
    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")

    ds.auto_update = True
    ds.hf_dataset_id = body.hf_dataset_id
    ds.hf_subset = body.hf_subset
    ds.hf_split = body.hf_split
    ds.update_interval_hours = body.update_interval_hours
    session.add(ds)
    await session.commit()
    await session.refresh(ds)
    return ds


@router.post("/{dataset_id}/unsubscribe", response_model=DatasetResponse)
async def unsubscribe_dataset(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
):
    """Disable auto-update subscription for a dataset."""
    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")

    ds.auto_update = False
    ds.sync_status = ""
    session.add(ds)
    await session.commit()
    await session.refresh(ds)
    return ds


@router.post("/{dataset_id}/sync", response_model=DatasetResponse)
async def sync_dataset_now(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
):
    """Manually trigger a sync check for a dataset."""
    from app.services.dataset_sync import check_and_sync_dataset

    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")

    if not ds.hf_dataset_id and ds.source_type not in (SourceType.huggingface, SourceType.preset):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "仅支持 HuggingFace/预设数据集的同步",
        )

    result = await check_and_sync_dataset(dataset_id, triggered_by="manual")
    if result.startswith("failed:"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, result[7:])

    await session.refresh(ds)
    return ds


@router.delete("/{dataset_id}", status_code=204)
async def delete_dataset(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
    storage: StorageBackend = Depends(_get_storage),
):
    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "数据集未找到")

    # Delete eval_results referencing this dataset
    stmt = select(EvalResult).where(EvalResult.dataset_id == dataset_id)
    results = (await session.exec(stmt)).all()
    for r in results:
        await session.delete(r)

    # Delete sync logs
    log_stmt = select(SyncLog).where(SyncLog.dataset_id == dataset_id)
    logs = (await session.exec(log_stmt)).all()
    for log in logs:
        await session.delete(log)

    # Delete dataset versions (must flush before deleting parent)
    await delete_dataset_versions(session, ds.id)
    await session.flush()

    await cleanup_uploaded_file(storage, ds)
    await session.delete(ds)
    await session.commit()


# ── Two-Stage Import: Preflight + Confirm ──────────────────────────


@router.post("/preflight", response_model=PreflightResponse)
async def preflight_import(
    file: UploadFile | None = None,
    source: str = Form(""),
    dataset_id: str = Form(""),
    subset: str = Form(""),
    split: str = Form("test"),
    server_path: str = Form(""),
    current_user: User = require_permission("datasets.write"),
    storage: StorageBackend = Depends(_get_storage),
):
    """Stage 1: Preview data, detect format, identify fields, validate.

    Supports all source types: upload file, HuggingFace, ModelScope, server path.
    Returns a preflight_token to use with /confirm.
    """
    _cleanup_preflight_cache()
    if len(_preflight_cache) >= _PREFLIGHT_MAX_SIZE:
        raise HTTPException(429, "预检操作过多，请稍后再试")

    from app.services.dataset_stats import _infer_dtype, _load_dataframe

    warnings: list[str] = []
    source_uri = ""
    fmt = ""
    size_bytes = 0

    if file and file.filename:
        # Upload source
        ext = os.path.splitext(file.filename)[1].lower()
        fmt = ext.lstrip(".") or "jsonl"
        content = await file.read()
        size_bytes = len(content)
        file_id = uuid.uuid4()
        key = f"preflight/{file_id}{ext}"
        source_uri = await storage.write_file(key, content)
        source_type = "upload"

    elif server_path:
        # Server path source
        if not os.path.exists(server_path):
            raise HTTPException(400, f"服务器路径不存在: {server_path}")
        source_uri = server_path
        ext = os.path.splitext(server_path)[1].lower()
        fmt = ext.lstrip(".") or "jsonl"
        size_bytes = os.path.getsize(server_path)
        source_type = "server_path"

    elif source in ("huggingface", "modelscope") and dataset_id:
        # Online source — download to staging area
        from app.services.dataset_import import import_huggingface, import_modelscope

        try:
            if source == "huggingface":
                source_uri, row_count, size_bytes = await import_huggingface(
                    dataset_id,
                    subset,
                    split,
                    storage,
                    hf_token=current_user.hf_token or None,
                )
            else:
                source_uri, row_count, size_bytes = await import_modelscope(
                    dataset_id,
                    subset,
                    split,
                    storage,
                    ms_token=current_user.ms_token or None,
                )
        except Exception as e:
            raise HTTPException(400, f"预检下载失败: {e}") from e
        ext = os.path.splitext(source_uri)[1].lower()
        fmt = ext.lstrip(".") or "jsonl"
        source_type = source
    else:
        raise HTTPException(400, "请提供文件、服务器路径或数据源 ID")

    # Load sample data
    try:
        df = await _load_dataframe(storage, source_uri)
    except Exception as e:
        raise HTTPException(400, f"文件解析失败: {e}") from e

    row_count = len(df)
    columns = list(df.columns)
    sample_rows = df.head(10).fillna("").to_dict(orient="records")

    # Field type inference
    field_types = {}
    for col in df.columns:
        field_types[col] = _infer_dtype(df[col])

    # Validation warnings
    if row_count == 0:
        warnings.append("数据集为空（0行）")
    if row_count > 0:
        null_cols = [col for col in df.columns if df[col].isna().sum() / row_count > 0.5]
        if null_cols:
            warnings.append(f"以下字段空值率超过 50%: {', '.join(null_cols)}")

    # Generate preflight token and cache
    token = hashlib.sha256(f"{source_uri}:{time.time()}".encode()).hexdigest()[:32]
    _preflight_cache[token] = {
        "source_uri": source_uri,
        "source_type": source_type,
        "format": fmt,
        "row_count": row_count,
        "size_bytes": size_bytes,
        "hf_dataset_id": dataset_id if source in ("huggingface", "modelscope") else "",
        "hf_subset": subset,
        "hf_split": split,
        "created_at": time.time(),
    }

    return PreflightResponse(
        source_type=source_type,
        format=fmt,
        row_count=row_count,
        size_bytes=size_bytes,
        columns=columns,
        sample_rows=sample_rows,
        field_types=field_types,
        warnings=warnings,
        preflight_token=token,
    )


@router.post("/confirm", response_model=DatasetResponse, status_code=201)
async def confirm_import(
    body: PreflightConfirmRequest,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.write"),
):
    """Stage 2: Confirm preflight and commit dataset to database."""
    cached = _preflight_cache.pop(body.preflight_token, None)
    if not cached:
        raise HTTPException(400, "预检令牌无效或已过期")

    # Check expiry (30 minutes)
    if time.time() - cached["created_at"] > 1800:
        raise HTTPException(400, "预检令牌已过期（30 分钟有效期）")

    source_type_map = {
        "upload": SourceType.upload,
        "server_path": SourceType.server_path,
        "huggingface": SourceType.huggingface,
        "modelscope": SourceType.modelscope,
    }
    st = source_type_map.get(cached["source_type"], SourceType.upload)

    # Auto-version
    stmt = select(Dataset).where(Dataset.name == body.name).order_by(Dataset.version.desc())
    existing = (await session.exec(stmt)).first()
    version = (existing.version + 1) if existing else 1

    ds = Dataset(
        name=body.name,
        description=body.description,
        source_type=st,
        source_uri=cached["source_uri"],
        format=cached["format"],
        tags=body.tags,
        version=version,
        size_bytes=cached["size_bytes"],
        row_count=cached["row_count"],
        created_by=current_user.id,
        hf_dataset_id=cached.get("hf_dataset_id", ""),
        hf_subset=cached.get("hf_subset", ""),
        hf_split=cached.get("hf_split", ""),
    )
    session.add(ds)
    await session.commit()
    await session.refresh(ds)

    dv = DatasetVersion(
        dataset_id=ds.id,
        version=version,
        file_path=cached["source_uri"],
        changelog="Initial import" if version == 1 else f"Version {version}",
        row_count=cached["row_count"],
        size_bytes=cached["size_bytes"],
        format=cached["format"],
    )
    session.add(dv)
    await session.commit()
    await session.refresh(ds)
    return ds


# ── Version Management ─────────────────────────────────────────────


@router.get("/{dataset_id}/versions", response_model=list[DatasetVersionResponse])
async def list_versions(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
):
    """List all versions of a dataset, newest first."""
    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(404, "数据集未找到")
    stmt = (
        select(DatasetVersion)
        .where(DatasetVersion.dataset_id == dataset_id)
        .order_by(DatasetVersion.version.desc())
    )
    result = await session.exec(stmt)
    return result.all()


@router.get(
    "/{dataset_id}/versions/{version}/preview",
)
async def preview_version(
    dataset_id: uuid.UUID,
    version: int,
    limit: int = 50,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
    storage: StorageBackend = Depends(_get_storage),
):
    """Preview rows from a specific dataset version."""
    stmt = select(DatasetVersion).where(
        DatasetVersion.dataset_id == dataset_id,
        DatasetVersion.version == version,
    )
    dv = (await session.exec(stmt)).first()
    if not dv:
        raise HTTPException(404, "版本未找到")

    # Reuse the same preview logic but with the version's file_path
    from app.services.dataset_stats import _load_dataframe

    try:
        df = await _load_dataframe(storage, dv.file_path)
    except Exception:
        return {"rows": [], "total": 0}

    rows = df.head(limit).fillna("").to_dict(orient="records")
    return {"rows": rows, "total": len(df)}


# ── Dataset Statistics ─────────────────────────────────────────────


@router.get("/{dataset_id}/stats", response_model=DatasetStatsResponse)
async def dataset_stats(
    dataset_id: uuid.UUID,
    version: int | None = None,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
    storage: StorageBackend = Depends(_get_storage),
):
    """Compute statistical summary for a dataset (or specific version)."""
    from app.services.dataset_stats import compute_dataset_stats

    ds = await session.get(Dataset, dataset_id)
    if not ds:
        raise HTTPException(404, "数据集未找到")

    source_uri = ds.source_uri
    size_bytes = ds.size_bytes

    # If version specified, use that version's file
    if version is not None:
        stmt = select(DatasetVersion).where(
            DatasetVersion.dataset_id == dataset_id,
            DatasetVersion.version == version,
        )
        dv = (await session.exec(stmt)).first()
        if not dv:
            raise HTTPException(404, f"版本 {version} 未找到")
        source_uri = dv.file_path
        size_bytes = dv.size_bytes or ds.size_bytes

    if not source_uri:
        raise HTTPException(400, "数据集无内容")

    try:
        stats = await compute_dataset_stats(storage, source_uri, size_bytes)
    except FileNotFoundError as e:
        raise HTTPException(404, str(e)) from e

    return stats


# ── Sync Logs ──────────────────────────────────────────────────────


@router.get(
    "/{dataset_id}/sync-logs",
    response_model=list[SyncLogResponse],
)
async def list_sync_logs(
    dataset_id: uuid.UUID,
    limit: int = 20,
    session: AsyncSession = Depends(get_db),
    current_user: User = require_permission("datasets.read"),
):
    """List sync history for a dataset."""
    stmt = (
        select(SyncLog)
        .where(SyncLog.dataset_id == dataset_id)
        .order_by(SyncLog.created_at.desc())
        .limit(limit)
    )
    result = await session.exec(stmt)
    return result.all()
